from __future__ import annotations

import re
from collections import OrderedDict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.db import SessionLocal
from app.core.security import get_password_hash
from app.models import (
    ActiveChecklist,
    ChecklistAssignment,
    ChecklistTemplate,
    Object,
    ShiftSession,
    TemplateItem,
    User,
)


TRUNCATE_TABLES = [
    "approval_requests",
    "browser_fingerprints",
    "user_locations",
    "notifications",
    "push_subscriptions",
    "personal_checklist_items",
    "shift_sessions",
    "checklist_progress",
    "active_checklists",
    "checklist_assignments",
    "template_items",
    "checklist_templates",
    "objects",
    "users",
]

CHECKLIST_ITEMS = [
    ("Пропылесосить / подмести", "Зал / Офис", "vacuum", 15),
    ("Протереть столы и стулья", "Зал / Офис", "table", 10),
    ("Вынести мусор", "Зал / Офис", "trash_bin", 5),
    ("Протереть подоконники", "Зал / Офис", "window", 5),
    ("Помыть полы", "Зал / Офис", "mop", 15),
]

ADDRESS_COORDS: dict[str, tuple[float, float]] = {
    "ул. Туран 44Б": (51.113389, 71.402303),
    "ул. Мангилик Ел, 19": (51.114362, 71.432846),
    "пр. Аль-Фараби, 55": (51.179940, 71.470954),
    "ул. Кенесары, 78": (51.1627124, 71.4626019),
    "ул. Кажымукана, 2": (51.1515051, 71.4590854),
    "ул. Женис 25": (51.170542, 71.412086),
    "ул. Сыганак, 10": (51.1241005, 71.4196463),
    "ул. Абылай хана 34": (51.153633, 71.485402),
    "ул. Кунаева 14": (51.1284264, 71.435346),
    "ул. Достык, 12": (51.1260853, 71.4261403),
    "ул. Валиханова 13": (51.168165, 71.439098),
    "ул. Мухаметханова, 11/1": (51.139217, 71.396661),
    "ул. Есенберлина, 12": (51.1908509, 71.4077546),
    "ул. Молдагулова 27": (51.1860665, 71.4104256),
    "БЦ Арман, архив": (51.160206, 71.411834),
    "БЦ Арман, региональная администрация": (51.160206, 71.411834),
    "БЦ Арман, отдел кадров": (51.160206, 71.411834),
    "ул. Туран 55": (51.112763, 71.398376),
    "ул. Мангилик Ел 42": (51.098986, 71.428970),
    "ул. Иманова 8": (51.1626871, 71.4302624),
}


def normalize_phone(value: object) -> str | None:
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        return None
    if len(digits) > 11:
        digits = digits[-11:]
    if len(digits) == 10:
        digits = f"7{digits}"
    elif len(digits) == 11 and digits.startswith("8"):
        digits = f"7{digits[1:]}"
    if len(digits) != 11 or not digits.startswith("7"):
        return None
    return f"+7 {digits[1:4]} {digits[4:7]} {digits[7:9]} {digits[9:11]}"


def normalize_iin(value: object) -> str | None:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) < 12:
        return None
    return digits[:12]


def normalize_name(value: object) -> str | None:
    if value is None:
        return None
    name = re.sub(r"\s+", " ", str(value)).strip(" .;")
    return name or None


def split_names(value: object) -> list[str]:
    if value is None:
        return []
    raw = str(value).replace("\n", " ")
    chunks = [re.sub(r"\s+", " ", part).strip(" .;") for part in re.split(r" {2,}|\t+", raw) if part and part.strip()]
    if len(chunks) > 1:
        return chunks
    merged = normalize_name(raw)
    if not merged:
        return []
    tokens = merged.split()
    if len(tokens) >= 4 and len(tokens) % 2 == 0:
        return [" ".join(tokens[i : i + 2]) for i in range(0, len(tokens), 2)]
    return [merged]


def split_digits(value: object) -> list[str]:
    if value is None:
        return []
    return re.findall(r"\d{10,13}", str(value))


def clean_address(value: object) -> str:
    if value is None:
        return ""
    text_value = str(value).replace('"', "").strip()
    text_value = re.sub(r"(?i)kaspi\s*bank", "", text_value)
    text_value = re.sub(r"(?i)\bao\b", "", text_value)
    text_value = re.sub(r"(?i)\bг\.\s*астана\b[,.]?", "", text_value)
    text_value = re.sub(r"(?i)\bастана\b[,.]?", "", text_value)
    text_value = re.sub(r"\s+", " ", text_value)
    text_value = text_value.strip(" ,.")
    text_value = text_value.replace("Альфараби", "Аль-Фараби")
    text_value = text_value.replace(" ,", ",")
    text_value = re.sub(r"\s*,\s*", ", ", text_value)

    lower = text_value.lower()
    known_prefixes = ("ул.", "пр.", "просп.", "мкр.", "жк.", "пер.", "наб.", "шоссе", "бц ", "тц ", "трц ")
    if text_value and not lower.startswith(known_prefixes):
        text_value = f"ул. {text_value}"
    return text_value


def load_rows(excel_path: Path) -> tuple[str, dict[str, str], list[dict]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    partner_name = ""
    curator_info: dict[str, str] = {}
    object_rows: list[dict] = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue

        partner_name = partner_name or (normalize_name(row[4]) or "")
        curator_info = curator_info or {
            "name": normalize_name(row[5]) or "Куратор",
            "iin": normalize_iin(row[6]) or "",
            "phone": normalize_phone(row[7]) or "",
        }

        address = clean_address(row[2])
        city = normalize_name(row[3]) or "Астана"
        cleaner_names = split_names(row[8])
        cleaner_iins = [normalize_iin(item) for item in split_digits(row[9])]
        cleaner_phones = [normalize_phone(item) for item in split_digits(row[10])]

        cleaners: list[dict[str, str]] = []
        count = min(len(cleaner_names), len(cleaner_iins), len(cleaner_phones))
        for idx in range(count):
            name = cleaner_names[idx]
            iin = cleaner_iins[idx]
            phone = cleaner_phones[idx]
            if not (name and iin and phone):
                continue
            cleaners.append({"name": name, "iin": iin, "phone": phone})

        if not address:
            continue

        object_rows.append(
            {
                "name": address,
                "address": address,
                "city": city,
                "cleaners": cleaners,
            }
        )

    return partner_name, curator_info, object_rows


def truncate_database(db: Session) -> None:
    joined = ", ".join(TRUNCATE_TABLES)
    db.execute(text(f"TRUNCATE TABLE {joined} RESTART IDENTITY CASCADE"))
    db.commit()


def import_kaspi_excel(excel_path: Path) -> dict[str, int]:
    partner_name, curator_info, object_rows = load_rows(excel_path)
    if not object_rows:
        raise RuntimeError("Excel file does not contain importable rows.")

    db = SessionLocal()
    try:
        truncate_database(db)

        admin = User(
            phone="+7 700 000 00 01",
            iin="111111111111",
            name="Админ Алина",
            role="admin",
            password_hash=get_password_hash("admin123"),
            status="active",
        )
        partner = User(
            phone="+7 700 000 00 02",
            iin="222222222222",
            name=partner_name or "Партнер",
            role="partner",
            password_hash=get_password_hash("partner123"),
            status="active",
        )
        real_curator = User(
            phone=curator_info["phone"],
            iin=curator_info["iin"],
            name=curator_info["name"],
            role="curator",
            parent=partner,
            password_hash=None,
            status="active",
        )
        test_curator = User(
            phone="+7 700 000 00 03",
            iin="333333333333",
            name=f"{curator_info['name']} (тест)",
            role="curator",
            parent=partner,
            password_hash=get_password_hash("curator123"),
            status="active",
        )
        test_cleaner = User(
            phone="+7 700 000 00 04",
            iin="444444444444",
            name="Клинер Дана",
            role="cleaner",
            parent=real_curator,
            password_hash=get_password_hash("cleaner123"),
            status="active",
        )

        db.add_all([admin, partner, real_curator, test_curator, test_cleaner])
        db.flush()

        cleaner_users: OrderedDict[tuple[str, str], User] = OrderedDict()
        for object_row in object_rows:
            for cleaner in object_row["cleaners"]:
                key = (cleaner["phone"], cleaner["iin"])
                if key in cleaner_users:
                    continue
                cleaner_users[key] = User(
                    phone=cleaner["phone"],
                    iin=cleaner["iin"],
                    name=cleaner["name"],
                    role="cleaner",
                    parent=real_curator,
                    password_hash=None,
                    status="active",
                )

        db.add_all(list(cleaner_users.values()))
        db.flush()

        template = ChecklistTemplate(
            name="Kaspi Bank",
            description="Базовый чек-лист для объектов Kaspi",
            created_by=admin.id,
            is_active=True,
        )
        db.add(template)
        db.flush()

        for index, (title, zone, icon, duration) in enumerate(CHECKLIST_ITEMS):
            db.add(
                TemplateItem(
                    template_id=template.id,
                    title=title,
                    zone=zone,
                    icon=icon,
                    duration_minutes=duration,
                    requires_photo=True,
                    order_index=index,
                )
            )
        db.flush()

        db.add(
            ChecklistAssignment(
                template_id=template.id,
                object_id=None,
                is_default=True,
                assigned_by=admin.id,
            )
        )
        db.flush()

        today = date.today()
        due_date = datetime.utcnow() + timedelta(hours=8)
        first_object_id: int | None = None

        for object_row in object_rows:
            coords = ADDRESS_COORDS.get(object_row["address"])
            obj = Object(
                name=object_row["name"],
                address=object_row["address"],
                city=object_row["city"],
                district=None,
                latitude=coords[0] if coords else None,
                longitude=coords[1] if coords else None,
                workers_count=max(len(object_row["cleaners"]), 1),
                status="active",
                partner_id=partner.id,
                curator_id=real_curator.id,
                created_by=admin.id,
            )
            db.add(obj)
            db.flush()
            if first_object_id is None:
                first_object_id = obj.id

            db.add(
                ChecklistAssignment(
                    template_id=template.id,
                    object_id=obj.id,
                    is_default=False,
                    assigned_by=admin.id,
                )
            )

            for cleaner in object_row["cleaners"]:
                key = (cleaner["phone"], cleaner["iin"])
                cleaner_user = cleaner_users.get(key)
                if not cleaner_user:
                    continue
                active = ActiveChecklist(
                    template_id=template.id,
                    object_id=obj.id,
                    assigned_to=cleaner_user.id,
                    assigned_by=real_curator.id,
                    status="pending",
                    due_date=due_date,
                    shift_date=today,
                    priority="normal",
                    notes="Импортировано из Excel",
                )
                db.add(active)
                db.flush()
                db.add(
                    ShiftSession(
                        user_id=cleaner_user.id,
                        object_id=obj.id,
                        active_checklist_id=active.id,
                        status="planned",
                    )
                )

        if first_object_id is not None:
            active = ActiveChecklist(
                template_id=template.id,
                object_id=first_object_id,
                assigned_to=test_cleaner.id,
                assigned_by=test_curator.id,
                status="pending",
                due_date=due_date,
                shift_date=today,
                priority="normal",
                notes="Тестовый доступ",
            )
            db.add(active)
            db.flush()
            db.add(
                ShiftSession(
                    user_id=test_cleaner.id,
                    object_id=first_object_id,
                    active_checklist_id=active.id,
                    status="planned",
                )
            )

        db.commit()

        return {
            "objects": len(object_rows),
            "real_cleaners": len(cleaner_users),
            "users_total": db.query(User).count(),
        }
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        raise SystemExit("Usage: python -m app.services.import_kaspi_excel <excel-path>")

    result = import_kaspi_excel(Path(sys.argv[1]))
    print(result)
